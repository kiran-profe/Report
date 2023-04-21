from __future__ import unicode_literals
from django.shortcuts import render, get_object_or_404
from web.forms import FileForm
from web.models import Report, FitnessReport
from django.db.models import Q
from django.urls import reverse
import datetime
from openpyxl import load_workbook
from django.http import JsonResponse
import openpyxl


def index(request):
    reports = FitnessReport.objects.filter(is_deleted=True)

    date = datetime.date.today()
    start_week = date - datetime.timedelta(date.weekday())
    end_week = start_week + datetime.timedelta(5)
    sunday = date - datetime.timedelta(days=date.weekday() + 1)

    tue = start_week + datetime.timedelta(1)
    wed = start_week + datetime.timedelta(2)
    thu = start_week + datetime.timedelta(3)
    fri = start_week + datetime.timedelta(4)

    if FitnessReport.objects.filter(date=sunday, is_deleted=True).exists():
        sunday_report = get_object_or_404(
            FitnessReport.objects.filter(date=sunday, is_deleted=True))

        srun = sunday_report.running_time
        gsrun = sunday_report.jogging_time

        esrun = sunday_report.exercise_time
        sun_tot = float(srun) + float(gsrun) + float(esrun)
        rensun = float(srun) / float(sun_tot)
        jonsun = float(gsrun) / float(sun_tot)
        exnsun = float(esrun) / float(sun_tot)
    else:
        sunday_report = 0
        srun = 0
        gsrun = 0
        esrun = 0
        sun_tot = 0
        rensun = 0
        jonsun = 0
        exnsun = 0

    if FitnessReport.objects.filter(date=start_week, is_deleted=True).exists():
        mon_report = get_object_or_404(
            FitnessReport.objects.filter(date=start_week, is_deleted=True))
        mrun = mon_report.running_time
        gmrun = mon_report.jogging_time
        emrun = mon_report.exercise_time

        mon_tot = float(mrun) + float(gmrun) + float(emrun)

        renmon = float(mrun) / float(mon_tot)
        jonmon = float(gmrun) / float(mon_tot)
        exnmon = float(emrun) / float(mon_tot)
    else:
        mon_report = 0
        mrun = 0
        gmrun = 0
        emrun = 0
        mon_tot = 0
        renmon = 0
        jonmon = 0
        exnmon = 0

    if FitnessReport.objects.filter(date=tue, is_deleted=True).exists():
        tue_report = get_object_or_404(
            FitnessReport.objects.filter(date=tue, is_deleted=True))
        turun = tue_report.running_time
        ghturun = tue_report.jogging_time
        ehturun = tue_report.jogging_time
        tue_tot = float(turun) + float(ghturun) + float(ehturun)

        rentu = float(turun) / float(tue_tot)
        jontu = float(ghturun) / float(tue_tot)
        exntu = float(ehturun) / float(tue_tot)
    else:
        tue_report = 0
        turun = 0
        ghturun = 0
        ehturun = 0
        tue_tot = 0
        rentu = 0
        jontu = 0
        exntu = 0

    if FitnessReport.objects.filter(date=wed, is_deleted=True).exists():
        wed_report = get_object_or_404(
            FitnessReport.objects.filter(date=wed, is_deleted=True))
        wrun = wed_report.running_time
        gwrun = wed_report.jogging_time
        ewrun = wed_report.exercise_time

        wed_tot = float(wrun) + float(gwrun) + float(ewrun)

        renwe = float(wrun) / float(wed_tot)
        jonwe = float(gwrun) / float(wed_tot)
        exnwe = float(ewrun) / float(wed_tot)
    else:
        wed_report = 0
        wrun = 0
        gwrun = 0
        ewrun = 0
        wed_tot = 0
        renwe = 0
        jonwe = 0
        exnwe = 0

    if FitnessReport.objects.filter(date=thu, is_deleted=True).exists():
        thu_report = get_object_or_404(
            FitnessReport.objects.filter(date=thu, is_deleted=True))
        thrun = thu_report.running_time
        gturun = tue_report.jogging_time
        eturun = tue_report.exercise_time
        thu_tot = float(thrun) + float(gturun) + float(eturun)

        renthu = float(thrun) / float(thu_tot)
        jonthu = float(gturun) / float(thu_tot)
        exnthu = float(eturun) / float(thu_tot)
    else:
        thu_report = 0
        thrun = 0
        gturun = 0
        eturun = 0
        thu_tot = 0
        renthu = 0
        jonthu = 0
        exnthu = 0

    if FitnessReport.objects.filter(date=fri, is_deleted=True).exists():
        fri_report = get_object_or_404(
            FitnessReport.objects.filter(date=fri, is_deleted=True))
        frun = fri_report.running_time
        gfrun = fri_report.jogging_time
        efrun = fri_report.exercise_time
        fri_tot = float(frun) + float(gfrun) + float(efrun)

        renfri = float(frun) / float(fri_tot)
        jonfri = float(gfrun) / float(fri_tot)
        exnfri = float(efrun) / float(fri_tot)
    else:
        fri_report = 0
        frun = 0
        gfrun = 0
        efrun = 0
        fri_tot = 0
        renfri = 0
        jonfri = 0
        exnfri = 0

    if FitnessReport.objects.filter(date=end_week, is_deleted=True).exists():
        sat_report = get_object_or_404(
            FitnessReport.objects.filter(date=end_week, is_deleted=True))
        sarun = sat_report.running_time
        gsarun = sat_report.jogging_time
        esarun = sat_report.exercise_time
        sat_tot = float(sarun) + float(gsarun) + float(esarun)

        rensat = float(sarun) / float(sat_tot)
        jonsat = float(gsarun) / float(sat_tot)
        exnsat = float(esarun) / float(sat_tot)
    else:
        sat_report = 0
        sarun = 0
        gsarun = 0
        esarun = 0
        sat_tot = 0
        rensat = 0
        jonsat = 0
        exnsat = 0

    runn_tot = float(srun) + \
        float(mrun) + float(turun) + \
        float(wrun) + float(thrun) + \
        float(frun) + float(sarun)
    jogg_tot = float(gsrun) + \
        float(gmrun) + float(gturun) + \
        float(gwrun) + float(ghturun) + \
        float(gfrun) + float(gsarun)
    ex_tot = float(esrun) + \
        float(emrun) + float(eturun) + \
        float(ewrun) + float(ehturun) + \
        float(efrun) + float(esarun)

    all_tot = float(runn_tot) + float(jogg_tot) + float(ex_tot)

    tot_runn_all = float(runn_tot) / float(all_tot)
    tot_jogg_all = float(jogg_tot) / float(all_tot)
    tot_ex_all = float(ex_tot) / float(all_tot)
    context = {
        "reports": reports,
        "sunday_report": sunday_report,
        "mon_report": mon_report,
        "tue_report": tue_report,
        "wed_report": wed_report,
        "thu_report": thu_report,
        "fri_report": fri_report,
        "sat_report": sat_report,
        "sunday": sunday,
        "monday": start_week,
        "tue": tue,
        "wed": wed,
        "thu": thu,
        "fri": fri,
        "sat": end_week,
        "runn_tot": runn_tot,
        "jogg_tot": jogg_tot,
        "ex_tot": ex_tot,
        "all_tot": all_tot,
        "sun_tot": sun_tot,
        "mon_tot": mon_tot,
        "tue_tot": tue_tot,
        "wed_tot": wed_tot,
        "thu_tot": thu_tot,
        "fri_tot": fri_tot,
        "sat_tot": sat_tot,
        "rensun": rensun,
        "jonsun": jonsun,
        "exnsun": exnsun,
        "renmon": renmon,
        "jonmon": jonmon,
        "exnmon": exnmon,
        "rentu": rentu,
        "jontu": jontu,
        "exntu": exntu,
        "renwe": renwe,
        "jonwe": jonwe,
        "exnwe": exnwe,
        "renthu": renthu,
        "jonthu": jonthu,
        "exnthu": exnthu,
        "renfri": renfri,
        "jonfri": jonfri,
        "exnfri": exnfri,
        "rensat": rensat,
        "jonsat": jonsat,
        "exnsat": exnsat,
        "tot_runn_all": tot_runn_all,
        "tot_jogg_all": tot_jogg_all,
        "tot_ex_all": tot_ex_all
    }
    return render(request, 'web/index.html', context)


def upload_data(request):
    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)

        if form.is_valid():
            input_excel = request.FILES['file']

            # book = xlrd.open_workbook(file_contents=input_excel.read())
            # sheet = book.sheet_by_index(0)
            # data = pd.read_excel(input_excel)
            wb = load_workbook(filename=input_excel, read_only=True)
            ws = wb['Sheet2']
            print("kk")
            print(ws.max_column)
            row_count = (ws.max_row) + 1
            column_count = (ws.max_column) + 1
            column_real = ws.max_column
            dict_list = []

            # for row in wb.rows:
            c = 0
            keys = [str(ws.cell(1, col_index).value)
                    for col_index in range(1, column_count)]
            for row_index in range(1, row_count):
                c = c+1
                d = {keys[col_index]: str(ws.cell(row_index, (col_index + 1)).value)
                     for col_index in range(column_real)}

                dict_list.append(d)
                # for cell in row:
                #     print("sheet")
                #     print(cell.value)

            # dict_list = []
            # keys = [str(sheet.cell(0, col_index).value) for col_index in range(sheet.ncols)]
            # for row_index in range(1, sheet.nrows):
            #     d = {keys[col_index]: str(sheet.cell(row_index, col_index).value)
            #         for col_index in range(sheet.ncols)}
            #     dict_list.append(d)

            is_ok = True
            message = ''
            row_count = 2
            for item in dict_list:
                date = item['Date']
                jogging = item['Jogging']
                running = item['Running']
                exercise = item['Exercise']

                database = FitnessReport.objects.create(
                    date=date,
                    jogging_time=jogging,
                    running_time=running,
                    exercise_time=exercise
                )
                database.save()

            response_data = {
                "status": "true",
                "title": "Successfully Updated",
                "message": "Customer Updated successfully.",
                "redirect": "true",
                "redirect_url": reverse('web:upload_data')
            }

        else:
            message = print(form.errors)
            x = message[0]
            form = FileForm()
            response_data = {
                'status': 'false',
                'stable': 'true',
                'title': "Form validation error",
                "message": str(x),
                "form": form,
            }
        return JsonResponse(response_data)
    else:
        form = FileForm()

        context = {
            "form": form,
            "title": "Upload",
            "caption": "Upload",
            "redirect": True,
            "url": reverse('web:upload_data'),
        }
        return render(request, 'web/upload.html', context)
